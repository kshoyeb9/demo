"""
etl_pl.py — Rebuilds the Noon P&L dashboard from the input workbook.

Usage:
    python etl_pl.py --file noon_dashboard_input.xlsx [--out index.html]

The twelve-month grids on the 'Dashboard Input' tab are the single source of
truth. The workbook's own month/YTD cells are formulas rolling up from those
grids; this script deliberately does NOT read them — openpyxl returns None for
a formula Excel has not cached, so anything read that way would silently be
zero. Instead it re-derives the same figures from the same grids, cut off at
the active month. Excel and the dashboard therefore agree whether or not the
workbook has ever been opened.

Tables are found by header text, so rows can be inserted or removed freely.
"""

import argparse, json, re, sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

ap = argparse.ArgumentParser()
ap.add_argument("--file", default="noon_dashboard_input.xlsx")
ap.add_argument("--out",  default="index.html")
ap.add_argument("--template", default=None, help="HTML template (defaults to --out)")
ap.add_argument("--publish-out", default="dashboard.publish.html",
                help="Also write a bare-content copy for publishing as an Artifact")
args = ap.parse_args()

xl_path, out_path = Path(args.file), Path(args.out)
tpl_path = Path(args.template) if args.template else out_path
if not xl_path.exists():  sys.exit(f"Workbook not found: {xl_path}")
if not tpl_path.exists(): sys.exit(f"HTML template not found: {tpl_path}")

wb = openpyxl.load_workbook(xl_path, data_only=True)
DATA_SHEET = "Dashboard Input"

def sheet(name=DATA_SHEET):
    if name in wb.sheetnames: return wb[name]
    cand = [n for n in wb.sheetnames if n.strip().lower() != "instructions"]
    if len(cand) == 1: return wb[cand[0]]
    sys.exit(f"Sheet '{name}' missing from {xl_path.name}. Found: {', '.join(wb.sheetnames)}")

ws   = sheet()
rows = list(ws.iter_rows(values_only=True))

def find_header(header, col=0):
    want = str(header).strip().lower()
    for i, r in enumerate(rows):
        if col < len(r) and r[col] is not None and str(r[col]).strip().lower() == want:
            return i
    return None

def read_table(header, ncols, *, col=0, stop_on_total=True, stop_labels=()):
    """Rows under `header` until a blank first cell. Total rows are dropped —
    they are recomputed here. `stop_labels` is per-call on purpose: a label that
    ends one table can be a valid data row in another."""
    h = find_header(header, col)
    if h is None:
        sys.exit(f"Header '{header}' not found on '{ws.title}'. "
                 f"Do not rename or delete a table's header row.")
    extra = {s.strip().lower() for s in stop_labels}
    out = []
    for r in rows[h+1:]:
        first = r[col] if col < len(r) else None
        if first is None or str(first).strip() == "": break
        lab = str(first).strip().lower()
        if lab in extra: break
        if stop_on_total and (lab.startswith("total") or lab.startswith("ytd total")): break
        out.append(tuple(r[i] if i < len(r) else None for i in range(col, col+ncols)))
    return out

def num(x, default=0.0):
    if isinstance(x, (int, float)): return float(x)
    if isinstance(x, str):
        try: return float(x.strip().replace("$","").replace(",","").replace("%",""))
        except ValueError: return default
    return default

def has(x):
    """True when a cell holds a real number (blank future months must not read 0)."""
    return isinstance(x, (int, float))

def yes(x): return str(x).strip().lower() in ("y","yes","true","1")
def rnd(x, d=4): return None if x is None else round(float(x), d)

def blk(a, b):
    a, b = float(a or 0), float(b or 0)
    return {"actual": rnd(a), "budget": rnd(b), "var": rnd(a-b), "pct": rnd(a/b) if b else None}


# ══ SETUP ═════════════════════════════════════════════════════════════════════
def setting(label, default=""):
    want = str(label).strip().lower()
    for r in rows:
        if r and r[0] is not None and str(r[0]).strip().lower() == want:
            return r[1] if len(r) > 1 and r[1] is not None else default
    return default

period_full = str(setting("Period label (month)", "Unknown"))
ytd_label   = str(setting("YTD label", ""))
year        = str(setting("Reporting year", datetime.now().year)).strip()
fiscal_year = str(setting("Fiscal year label", ""))
fy_start    = int(num(setting("Fiscal year start month", 7), 7))
currency    = str(setting("Currency label", "USD M"))
ACT         = int(num(setting("Active month number", 0), 0))

all_month_rows = read_table("Setup Month", 2)
all_months = [str(r[1]).strip() for r in all_month_rows if r[1]]
if not all_months:
    sys.exit("No months listed under 'Setup Month' on the Dashboard Input tab.")
if not 1 <= ACT <= len(all_months):
    sys.exit(f"'Active month number' is {ACT or 'blank'}; it must be between 1 and "
             f"{len(all_months)} (1 = {all_months[0]} … {len(all_months)} = {all_months[-1]}).")

months = all_months[:ACT]          # closed months only — this drives the period selector
NM     = ACT

meta = {"period": period_full, "ytd_label": ytd_label,
        "generated": datetime.now().strftime("%d %b %Y"), "source": xl_path.name,
        "currency": currency, "months": months, "year": year,
        "period_index": NM-1, "fy_start_month": fy_start, "fiscal_year": fiscal_year}


# ══ MONTHLY GRIDS — the source everything else is derived from ════════════════
NCOL = len(all_months)

def grid(header):
    """{line name: [12 raw cells]} for one monthly block."""
    return {str(r[0]).strip(): [r[1+i] for i in range(NCOL)]
            for r in read_table(header, 1+NCOL) if r[0]}

rev_act, rev_bgt = grid("Revenue — Actual"), grid("Revenue — Budget")
cost_act, cost_bgt = grid("Costs — Actual"), grid("Costs — Budget")

def series(g, name):            # closed months only, blanks as 0
    return [rnd(num(v)) for v in g.get(name, [])[:NM]]

def at_active(g, name):
    s = g.get(name, [])
    return num(s[NM-1]) if len(s) >= NM else 0.0

def upto_active(g, name):
    return sum(num(v) for v in g.get(name, [])[:NM])

def line_items(act_g, bgt_g, order):
    out = []
    for n in order:
        out.append({"name": n,
                    "month": blk(at_active(act_g, n),  at_active(bgt_g, n)),
                    "ytd":   blk(upto_active(act_g, n), upto_active(bgt_g, n)),
                    "total": False,
                    "monthly": {"actual": series(act_g, n), "budget": series(bgt_g, n)}})
    return out

rev_names  = [n for n in rev_act  if n.lower() != "total"]
cost_names = [n for n in cost_act if n.lower() != "total"]
revenue = line_items(rev_act, rev_bgt, rev_names)
costs   = line_items(cost_act, cost_bgt, cost_names)

def add_total(items, label):
    am = sum(i["month"]["actual"] for i in items); bm = sum(i["month"]["budget"] for i in items)
    ay = sum(i["ytd"]["actual"]   for i in items); by = sum(i["ytd"]["budget"]   for i in items)
    mon = {k: [rnd(sum(i["monthly"][k][x] or 0 for i in items)) for x in range(NM)]
           for k in ("actual","budget")}
    items.append({"name": label, "month": blk(am,bm), "ytd": blk(ay,by),
                  "total": True, "monthly": mon})
    return items[-1]

rev_total  = add_total(revenue, "Total Revenue")
cost_total = add_total(costs,   "Total costs")


# ══ P&L ═══════════════════════════════════════════════════════════════════════
pl_src = [(str(r[0]).strip(), str(r[6] or "No").strip())
          for r in read_table("P&L Line", 7) if r[0]]
cost_by_name = {c["name"]: c for c in costs if not c["total"]}

def resolve_source(name):
    if name.lower() == "revenue": return rev_total
    if name in cost_by_name: return cost_by_name[name]
    stripped = re.sub(r"\s+(expense|costs?)$", "", name, flags=re.I).strip().lower()
    for cn, c in cost_by_name.items():
        if cn.lower() in (stripped, name.lower()): return c
    return None

def combine(base, minus, label):
    o = {"name": label, "kind": "value", "cost": False, "total": True}
    for p in ("month","ytd"):
        a = base[p]["actual"] - sum(m[p]["actual"] for m in minus)
        b = base[p]["budget"] - sum(m[p]["budget"] for m in minus)
        o[p] = blk(a,b)
    o["monthly"] = {k: [rnd(base["monthly"][k][i] - sum(m["monthly"][k][i] for m in minus))
                        for i in range(NM)] for k in ("actual","budget")}
    return o

pl, res = [], {}
for name, is_cost in pl_src:
    ln = name.lower()
    if ln == "gross profit":
        e = combine(res["Revenue"], [res["Direct costs"]], name)
    elif ln == "contribution profit":
        e = combine(res["Gross profit"], [res["Marketing expense"]], name)
    elif ln == "ebitda":
        opex = [res[n] for n in ("BU salaries","Noon HQ","Other operating expenses") if n in res]
        e = combine(res["Contribution profit"], opex, name)
    else:
        s = resolve_source(name)
        if s is None:
            sys.exit(f"P&L line '{name}' has no matching row in the revenue or cost grids.")
        e = {"name": name, "kind": "value", "month": dict(s["month"]), "ytd": dict(s["ytd"]),
             "monthly": s.get("monthly"), "cost": yes(is_cost), "total": False}
    res[name] = e; pl.append(e)

rev_e = res.get("Revenue", rev_total)
def margin(label, numer):
    e = {"name": label, "kind": "pct", "cost": False, "total": False, "monthly": None}
    for p in ("month","ytd"):
        a = (numer[p]["actual"] or 0)/(rev_e[p]["actual"] or 1)
        b = (numer[p]["budget"] or 0)/(rev_e[p]["budget"] or 1)
        e[p] = {"actual": rnd(a), "budget": rnd(b), "var": rnd(a-b), "pct": None}
    return e
for label, srcn in [("Gross profit margin","Gross profit"),
                    ("Contribution margin","Contribution profit"),
                    ("EBITDA margin","EBITDA")]:
    if srcn in res:
        i = next((k for k,p in enumerate(pl) if p["name"]==srcn), len(pl)-1)
        pl.insert(i+1, margin(label, res[srcn]))

def m_of(lbl,p):
    e = next((x for x in pl if x["name"]==lbl), None); return e[p]["actual"] if e else None
margins = {"gm_month": m_of("Gross profit margin","month"), "gm_ytd": m_of("Gross profit margin","ytd"),
           "ebitda_month": m_of("EBITDA margin","month"), "ebitda_ytd": m_of("EBITDA margin","ytd")}

tm, ty = rev_total["month"]["actual"] or 1, rev_total["ytd"]["actual"] or 1
rev_mix = [{"name": r["name"],
            "month": {"rev": r["month"]["actual"], "share": rnd(r["month"]["actual"]/tm)},
            "ytd":   {"rev": r["ytd"]["actual"],   "share": rnd(r["ytd"]["actual"]/ty)}}
           for r in revenue if not r["total"]]


# ══ WORKING CAPITAL ═══════════════════════════════════════════════════════════
wc_raw = read_table("WC Month", 8)[:NM]
wc_monthly, prev = [], None
for r in wc_raw:
    ar_,ap_,dfr,oth = num(r[1]),num(r[2]),num(r[3]),num(r[4])
    nwc = ar_+ap_+dfr+oth
    wc_monthly.append({"month": str(r[0]).strip(), "ar": rnd(ar_), "ap": rnd(ap_),
                       "deferred": rnd(dfr), "other": rnd(oth), "nwc": rnd(nwc),
                       "movement": rnd(0 if prev is None else nwc-prev)})
    prev = nwc

wc_ytd = []
if wc_monthly:
    o, c = wc_monthly[0], wc_monthly[-1]
    for item, key in [("Accounts receivable","ar"),("Accounts payable","ap"),
                      ("Deferred revenue","deferred"),("Other working capital items","other")]:
        wc_ytd.append({"item": item, "open": o[key], "close": c[key],
                       "movement": rnd(c[key]-o[key]), "cash_impact": rnd(-(c[key]-o[key]))})
    to_, tc = sum(x["open"] for x in wc_ytd), sum(x["close"] for x in wc_ytd)
    wc_ytd.append({"item": "Net working capital", "open": rnd(to_), "close": rnd(tc),
                   "movement": rnd(tc-to_), "cash_impact": rnd(-(tc-to_))})


# ══ AR / AP ═══════════════════════════════════════════════════════════════════
ar_monthly, opening = [], None
for r in read_table("AR Month", 6)[:NM]:
    inv, coll = num(r[2]), num(r[3])
    op = num(r[1]) if opening is None else opening
    cl = op + inv - coll
    ar_monthly.append({"month": str(r[0]).strip(), "opening": rnd(op), "invoiced": rnd(inv),
                       "collected": rnd(coll), "closing": rnd(cl),
                       "rate": rnd(coll/inv) if inv else None})
    opening = cl
inv_y  = sum(x["invoiced"]  for x in ar_monthly)
coll_y = sum(x["collected"] for x in ar_monthly)
ar_ytd = {"invoiced": rnd(inv_y), "collected": rnd(coll_y),
          "closing": ar_monthly[-1]["closing"] if ar_monthly else 0,
          "rate": rnd(coll_y/inv_y) if inv_y else None}

ap_monthly, opening = [], None
for r in read_table("AP Month", 6)[:NM]:
    pur, pay = num(r[2]), num(r[3])
    op = num(r[1]) if opening is None else opening
    cl = op + pur - pay
    ap_monthly.append({"month": str(r[0]).strip(), "opening": rnd(op), "purchases": rnd(pur),
                       "payments": rnd(pay), "closing": rnd(cl), "dpo": rnd(num(r[5]))})
    opening = cl
ap_ytd = {"purchases": rnd(sum(x["purchases"] for x in ap_monthly)),
          "payments":  rnd(sum(x["payments"]  for x in ap_monthly)),
          "closing":   ap_monthly[-1]["closing"] if ap_monthly else 0,
          "dpo":       rnd(sum(x["dpo"] for x in ap_monthly)/len(ap_monthly)) if ap_monthly else 0}

def aging(header):
    src = [(str(r[0]).strip(), num(r[1])) for r in read_table(header, 3)]
    tot = rnd(sum(a for _,a in src))
    return [{"bucket": b, "amount": rnd(a), "share": rnd(a/tot) if tot else None}
            for b,a in src], tot
ar_aging, ar_aging_total = aging("AR Aging Bucket")
ap_aging, ap_aging_total = aging("AP Aging Bucket")
ar_by_contract = [{"contract": str(r[0]).strip(), "amount": rnd(num(r[1]))}
                  for r in read_table("Contract", 3)]
ap_by_vendor   = [{"vendor": str(r[0]).strip(), "amount": rnd(num(r[1]))}
                  for r in read_table("Vendor", 3)]


# ══ CASH — the whole section is derived from the monthly cash-flow grid ═══════
cf = {str(r[0]).strip(): [r[1+i] for i in range(NCOL)]
      for r in read_table("Cash Flow Line", 1+NCOL) if r[0]}
INFLOWS  = ["Collections", "Other inflows"]
OUTFLOWS = ["Operating costs", "Capex", "Debt service", "Other outflows"]
for need in ["Opening cash"] + INFLOWS + OUTFLOWS:
    if need not in cf:
        sys.exit(f"Row '{need}' is missing from the 'Cash Flow Line' grid in Section 9.")

def cfv(name, i): return num(cf[name][i]) if i < len(cf[name]) else 0.0

# Chain the ledger forward: only January's opening is typed, each later month
# opens where the previous one closed.
net, closing, opening_m = [], [], []
carry = num(cf["Opening cash"][0])
for i in range(NM):
    n = sum(cfv(k,i) for k in INFLOWS) - sum(cfv(k,i) for k in OUTFLOWS)
    opening_m.append(carry); net.append(n); carry += n; closing.append(carry)

cash_close = closing[-1] if closing else 0.0
cash_open  = opening_m[0] if opening_m else 0.0

def bridge(per_month):
    i = NM-1
    steps = [{"step": "Opening cash", "kind": "total",
              "value": rnd(opening_m[i] if per_month else cash_open)}]
    run = steps[0]["value"]
    for k in INFLOWS:
        v = cfv(k,i) if per_month else sum(cfv(k,x) for x in range(NM))
        run += v; steps.append({"step": k, "kind": "delta", "value": rnd(v), "running": rnd(run)})
    for k in OUTFLOWS:
        v = cfv(k,i) if per_month else sum(cfv(k,x) for x in range(NM))
        run -= v; steps.append({"step": k, "kind": "delta", "value": rnd(-v), "running": rnd(run)})
    steps.append({"step": "Closing cash", "kind": "total", "value": rnd(cash_close)})
    return steps
bridge_month, bridge_ytd = bridge(True), bridge(False)

burn_3m = sum(net[max(0,NM-3):NM]) / len(net[max(0,NM-3):NM]) if net else 0.0
runway = {"burn_month": rnd(net[-1] if net else 0), "burn_3m": rnd(burn_3m),
          "ytd_move":   rnd(sum(net)),
          "months":     rnd(cash_close/-burn_3m, 2) if burn_3m < 0 else None}

# Cash balance chart: prior-year rows are typed; this year's come off the ledger
month_ix = {m.lower(): i for i, m in enumerate(all_months)}
cm_labels, cm_values, cm_fc, cm_ill = [], [], [], []
for r in read_table("Cash Month", 4):
    label = str(r[0]).strip()
    key = label[:3].lower()
    i = month_ix.get(key)
    same_year = i is not None and label[-2:] == (all_months and str(all_month_rows[i][0])[-2:])
    cm_labels.append(label)
    if same_year:
        cm_values.append(rnd(closing[i], 3) if i < NM else None)
        cm_fc.append(i >= NM)
    else:
        cm_values.append(rnd(num(r[1]), 3)); cm_fc.append(yes(r[2]))
    cm_ill.append(yes(r[3]))
keep = [j for j,v in enumerate(cm_values) if v is not None]
cash_monthly = {"labels":[cm_labels[j] for j in keep], "values":[cm_values[j] for j in keep],
                "forecast":[cm_fc[j] for j in keep], "illustrative":[cm_ill[j] for j in keep]}

TILE_KEY = {"cash balance":"cash", "collections":"collections",
            "net working capital":"nwc", "accounts receivable":"ar"}
tile_note = {}
for r in read_table("Cash Tile", 8):
    k = TILE_KEY.get(str(r[0]).strip().lower())
    if k: tile_note[k] = (str(r[2] or ""), str(r[6] or ""))
nwc_now = wc_monthly[-1]["nwc"] if wc_monthly else 0
nwc_open = wc_monthly[0]["nwc"] if wc_monthly else 0
tile_val = {
    "cash":        (cash_close, cash_close),
    "collections": (cfv("Collections", NM-1), sum(cfv("Collections",x) for x in range(NM))),
    "nwc":         (nwc_now, rnd(nwc_now - nwc_open)),
    "ar":          (ar_ytd["closing"], ar_ytd["closing"]),
}
cash_tiles = {"month": {}, "ytd": {}}
for k,(mv,yv) in tile_val.items():
    mn, yn = tile_note.get(k, ("",""))
    cash_tiles["month"][k] = {"value": rnd(mv), "note": mn}
    cash_tiles["ytd"][k]   = {"value": rnd(yv), "note": yn}
missing = set(TILE_KEY.values()) - set(cash_tiles["month"])
if missing:
    sys.exit(f"Cash tiles missing: {', '.join(sorted(missing))}. The 'Cash Tile' table "
             f"needs one row per tile with no blank rows between them.")

key_updates = [{"topic": str(r[1]).strip(), "text": str(r[2]).strip()}
               for r in read_table("Update #", 3) if r[1] and r[2]]


# ══ ASSEMBLE & INJECT ═════════════════════════════════════════════════════════
DATA = {"meta":meta, "revenue":revenue, "costs":costs, "pl":pl, "margins":margins,
        "rev_mix":rev_mix, "wc_monthly":wc_monthly, "wc_ytd":wc_ytd,
        "ar_monthly":ar_monthly, "ar_ytd":ar_ytd, "ar_aging":ar_aging,
        "ar_aging_total":ar_aging_total, "ar_by_contract":ar_by_contract,
        "ap_monthly":ap_monthly, "ap_ytd":ap_ytd, "ap_aging":ap_aging,
        "ap_aging_total":ap_aging_total, "ap_by_vendor":ap_by_vendor,
        "cash_tiles":cash_tiles, "bridge_month":bridge_month, "bridge_ytd":bridge_ytd,
        "cash_monthly":cash_monthly, "runway":runway, "key_updates":key_updates}

html = tpl_path.read_text(encoding="utf-8")
payload = json.dumps(DATA, ensure_ascii=False, separators=(",",":"))
new_html, n = re.subn(r"const DATA\s*=\s*\{.*?\};",
                      lambda _m: f"const DATA = {payload};", html, count=1, flags=re.DOTALL)
if n == 0: sys.exit("Could not find 'const DATA = {...};' in the template.")
out_path.write_text(new_html, encoding="utf-8")

if args.publish_out:
    try:
        s = new_html.index("<body>")+len("<body>"); e = new_html.rindex("</body>")
        Path(args.publish_out).write_text(new_html[s:e].strip()+"\n", encoding="utf-8")
    except ValueError:
        pass

def warn(m): print(f"  ! {m}")

# The likeliest mistake with an active-month pointer is moving it before the
# month's figures are in. Say so loudly rather than quietly reporting zeros.
if not any(has(v) for v in rev_act.get(rev_names[0], [])[NM-1:NM]) if rev_names else False:
    warn(f"NO ACTUALS for {all_months[ACT-1]} — the active month is set to {ACT} but the "
         f"Section 5 grids are blank there. Month figures will read zero. "
         f"Either enter the month, or set 'Active month number' back to {ACT-1}.")
print(f"✓ Rebuilt {out_path} from {xl_path.name}")
print(f"  Active    month {ACT} ({all_months[ACT-1]}) — rolling up {NM} month(s): {', '.join(months)}")
print(f"  Period    {meta['period']}   ·   YTD {meta['ytd_label']}")
print(f"  Revenue   ${rev_total['ytd']['actual']:.2f}M YTD"
      + (f" ({rev_total['ytd']['pct']*100:.1f}% of budget)" if rev_total['ytd']['pct'] else ""))
print(f"  EBITDA    {(margins['ebitda_ytd'] or 0)*100:.1f}% margin YTD")
print(f"  Cash      ${cash_close:.2f}M closing · {runway['months'] if runway['months'] is not None else '—'} mo runway")
print(f"  Tables    {len(revenue)-1} BUs · {len(costs)-1} costs · {len(pl)} P&L rows · "
      f"{len(ar_by_contract)} contracts · {len(ap_by_vendor)} vendors · "
      f"{len(cash_monthly['labels'])} cash months · {len(key_updates)} updates")

# Cross-checks the reader would otherwise have to eyeball
for r in revenue + costs:
    if r.get("monthly"):
        s = sum(r["monthly"]["actual"])
        if abs(s - r["ytd"]["actual"]) > 0.02:
            warn(f"'{r['name']}' monthly actuals sum to {s:.2f} but YTD reads {r['ytd']['actual']:.2f}")
if wc_monthly and abs(ar_ytd["closing"] - wc_monthly[-1]["ar"]) > 0.02:
    warn(f"AR closing {ar_ytd['closing']:.2f} does not match working-capital receivables "
         f"{wc_monthly[-1]['ar']:.2f} for {months[-1]}")
if abs(sum(net) - (cash_close - cash_open)) > 0.005:
    warn("cash-flow net movement does not reconcile to closing less opening cash")
for nm_, arr in (("AR", ar_monthly), ("AP", ap_monthly), ("working capital", wc_monthly)):
    if len(arr) < NM:
        warn(f"{nm_} has {len(arr)} month(s) of data but the active month is {NM}")
if not ar_by_contract: warn("AR by contract is empty — that chart will render blank.")
if not ap_by_vendor:   warn("AP by vendor is empty — that chart will render blank.")
