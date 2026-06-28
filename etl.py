"""etl.py — Noon Academy Financials ETL
Reads the financial model, CF tracker, and billing schedule files,
then writes fy_data.json.

Usage (run from F:\\Noon Dashboard\\):
    python etl.py --model "Noon financial model 2.0 v42.xlsx" ^
                  --cf    "Weekly Report CF 2026 Final.xlsx" ^
                  --schools-billing "Schools_Billing_Updated_May18.xlsx" ^
                  --b2b-billing     "B2B_Billing_Schedule_Updated_May18.xlsx"
"""

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run:  pip install openpyxl")

# ── Constants ─────────────────────────────────────────────────────────────────
SAR_TO_USD = 1 / 3.75   # billing files are in SAR; model is already USD

LABELS_2425 = ["Jul-24","Aug-24","Sep-24","Oct-24","Nov-24","Dec-24",
               "Jan-25","Feb-25","Mar-25","Apr-25","May-25","Jun-25"]
LABELS_2526 = ["Jul-25","Aug-25","Sep-25","Oct-25","Nov-25","Dec-25",
               "Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26"]

# First-of-month dates for FY25/26, used for contract overlap checks
FY2526_DATES = [datetime.strptime(lbl, "%b-%y") for lbl in LABELS_2526]

# ── BU tab configuration ──────────────────────────────────────────────────────
BU_TABS = {
    "Tracks":            ("03_BU_Tracks",            ["Net revenue",  "Net Revenue"]),
    "Government Schools":("04_BU_Government Schools", ["Net revenue",  "Net Revenue"]),
    "B2B":               ("05_BU_B2B",               ["Net Revenue (NGOs and MCIT)", "Net Revenue"]),
    "KSA B2C":           ("06_BU_KSA_B2C",           ["Total Revenue", "Net revenue"]),
    "Global Non-Saudi":  ("07_Global_Non_Saudi BU",  ["Total Revenue", "Net revenue"]),
    "Out of School":     ("06_BU_Out_of_School",     ["Net revenues",  "Net revenue"]),
}

GP_LABELS     = ["Gross profit", "Gross Margin", "Gross margin"]
CM_LABELS     = ["Contribution margin", "Contribution Margin"]
EBITDA_LABELS = ["EBITDA"]

FY2425_COLS = list(range(5, 17))   # cols E–P  (Jul-24 … Jun-25)
FY2526_COLS = list(range(18, 30))  # cols R–AC (Jul-25 … Jun-26)


# ── Shared helpers ────────────────────────────────────────────────────────────
def to_label(val) -> str | None:
    """Convert a date value (datetime, date, or string) to 'Mon-YY' label."""
    if isinstance(val, (datetime, date)):
        return val.strftime("%b-%y")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%b-%Y", "%B %Y", "%b %y", "%b-%y"):
        try:
            return datetime.strptime(s[:10] if fmt == "%Y-%m-%d" else s, fmt).strftime("%b-%y")
        except ValueError:
            pass
    return None


def to_dt(val) -> datetime | None:
    """Convert openpyxl date cell value to datetime."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None


# ── Financial model extraction ────────────────────────────────────────────────
def find_row(ws, labels: list, label_col: int = 3, max_row: int = 80) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=label_col, max_col=label_col):
        cell_val = row[0].value
        if cell_val is None:
            continue
        for lbl in labels:
            if str(cell_val).strip().lower() == lbl.lower():
                return row[0].row
    return None


def row_values(ws, row: int, cols: list) -> list:
    return [float(ws.cell(row=row, column=c).value or 0) for c in cols]


def extract_bu(ws, revenue_labels: list, cols: list) -> dict:
    rev_row = find_row(ws, revenue_labels)
    gp_row  = find_row(ws, GP_LABELS)
    cm_row  = find_row(ws, CM_LABELS)
    eb_row  = find_row(ws, EBITDA_LABELS)

    missing = [n for n, r in [("revenue", rev_row), ("gross_profit", gp_row),
               ("contribution_margin", cm_row), ("ebitda", eb_row)] if r is None]
    if missing:
        print(f"  WARNING: could not find rows for: {missing}")

    def safe(r):
        return row_values(ws, r, cols) if r else [0.0] * len(cols)

    return {
        "revenue":             safe(rev_row),
        "gross_profit":        safe(gp_row),
        "contribution_margin": safe(cm_row),
        "ebitda":              safe(eb_row),
    }


def build_fy_months(bu_data: dict, labels: list) -> list:
    months = []
    for i, lbl in enumerate(labels):
        by_bu = {bu: {f: round(bu_data[bu][f][i]) for f in bu_data[bu]} for bu in bu_data}
        months.append({"n": i + 1, "label": lbl, "by_bu": by_bu})
    return months


# ── CF tracker extraction ─────────────────────────────────────────────────────
def extract_cash(cf_path: Path) -> list:
    print(f"\nLoading CF tracker: {cf_path.name}")
    wb = openpyxl.load_workbook(cf_path, data_only=True)
    if "Monthly CF" not in wb.sheetnames:
        print("  WARNING: 'Monthly CF' sheet not found — cash will be zero.")
        return _empty_cash()
    ws = wb["Monthly CF"]

    # Row 5 = month headers; Row 94 = Ending cash
    cash_by_label = {}
    for col in range(2, ws.max_column + 1):
        hdr = ws.cell(row=5, column=col).value
        if hdr is None:
            continue
        lbl = to_label(hdr)
        if lbl not in LABELS_2526:
            continue
        eom = ws.cell(row=94, column=col).value or 0
        cash_by_label[lbl] = round(float(eom))

    result = []
    for i, lbl in enumerate(LABELS_2526):
        result.append({
            "n":         i + 1,
            "label":     lbl,
            "eom_cash":  cash_by_label.get(lbl, 0),
            "invoiced":  0,
            "collected": 0,
        })
    return result


def _empty_cash() -> list:
    return [{"n": i+1, "label": lbl, "eom_cash": 0, "invoiced": 0, "collected": 0}
            for i, lbl in enumerate(LABELS_2526)]


# ── Schools billing extraction ────────────────────────────────────────────────
def parse_schools_billing(path: Path) -> dict:
    """
    Returns {month_label: {'invoiced': USD, 'collected': USD}}.
    Sheets: Tanmiyah, Atyab, Olaya, Maarif — each has:
      Col B = invoice date, Col D = invoice amount (SAR)
      Col F = receipt date, Col H = receipt amount (SAR)
    """
    print(f"\nLoading schools billing: {path.name}")
    wb   = openpyxl.load_workbook(path, data_only=True)
    skip = {"Dashboard"}
    monthly: dict = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip:
            continue
        ws = wb[sheet_name]
        print(f"  Reading sheet: {sheet_name}")

        # Find the row that has "Date" in col B (the column header row)
        data_start = 7  # fallback
        for r in range(1, 15):
            val = ws.cell(row=r, column=2).value
            if val and str(val).strip() == "Date":
                data_start = r + 1
                break

        rows_read = 0
        for r in range(data_start, ws.max_row + 1):
            inv_date_val = ws.cell(row=r, column=2).value
            inv_amt_val  = ws.cell(row=r, column=4).value
            rec_date_val = ws.cell(row=r, column=6).value
            rec_amt_val  = ws.cell(row=r, column=8).value

            # Stop at a fully empty row (no date on either side)
            if inv_date_val is None and rec_date_val is None:
                break

            # Invoice
            if inv_date_val and inv_amt_val and str(inv_date_val).strip() not in ("TOTAL", ""):
                lbl = to_label(inv_date_val)
                if lbl:
                    monthly.setdefault(lbl, {"invoiced": 0.0, "collected": 0.0})
                    monthly[lbl]["invoiced"] += float(inv_amt_val) * SAR_TO_USD

            # Receipt
            if rec_date_val and rec_amt_val and str(rec_date_val).strip() not in ("TOTAL", ""):
                lbl = to_label(rec_date_val)
                if lbl:
                    monthly.setdefault(lbl, {"invoiced": 0.0, "collected": 0.0})
                    monthly[lbl]["collected"] += float(rec_amt_val) * SAR_TO_USD

            rows_read += 1

        print(f"    {rows_read} data rows processed")

    return monthly


# ── B2B billing extraction ────────────────────────────────────────────────────
def parse_b2b_billing(path: Path) -> dict:
    """
    Returns {month_label: {'invoiced': USD, 'collected': USD}}.
    Sheet 'B2B Billing Schedule', row 1 = header, data from row 2:
      Col 1 = Entity name
      Col 3 = Sign date
      Col 4 = End date
      Col 6 = Total contract value with VAT (SAR)
      Col 7 = Collected to date (SAR)
      Col 8 = Fraction of contract period elapsed ('How much passed?')
    Strategy: total_invoiced = Total × passed fraction, spread evenly across
    FY25/26 months within the contract period. Same for collected.
    """
    print(f"\nLoading B2B billing: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = "B2B Billing Schedule"
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not found — B2B billing skipped.")
        return {}
    ws = wb[sheet_name]

    monthly: dict = {}
    rows_read = 0

    for r in range(2, ws.max_row + 1):
        entity    = ws.cell(row=r, column=1).value
        if entity is None:
            break

        sign_val  = ws.cell(row=r, column=3).value
        end_val   = ws.cell(row=r, column=4).value
        total_vat = ws.cell(row=r, column=6).value
        collected = ws.cell(row=r, column=7).value
        passed    = ws.cell(row=r, column=8).value

        if not total_vat:
            continue

        total_vat = float(total_vat)
        collected = float(collected or 0)
        passed    = float(passed or 0)

        total_invoiced = total_vat * passed   # earned/billed to date (SAR)

        sign_dt = to_dt(sign_val)
        end_dt  = to_dt(end_val)
        if sign_dt is None or end_dt is None:
            continue

        # Find which FY25/26 months fall within the contract period
        active_months = [
            lbl for lbl, d in zip(LABELS_2526, FY2526_DATES)
            if sign_dt.replace(day=1) <= d <= end_dt.replace(day=1)
        ]
        if not active_months:
            # Contract outside FY25/26 entirely — skip
            continue

        n = len(active_months)
        inv_per_month = total_invoiced / n * SAR_TO_USD
        col_per_month = collected      / n * SAR_TO_USD

        entity_short = str(entity)[:30]
        print(f"  {entity_short:30s}  invoiced=${total_invoiced*SAR_TO_USD:,.0f}  "
              f"collected=${collected*SAR_TO_USD:,.0f}  over {n} months")

        for lbl in active_months:
            monthly.setdefault(lbl, {"invoiced": 0.0, "collected": 0.0})
            monthly[lbl]["invoiced"]  += inv_per_month
            monthly[lbl]["collected"] += col_per_month

        rows_read += 1

    print(f"  {rows_read} contracts processed")
    return monthly


# ── Merge billing into cash array ─────────────────────────────────────────────
def merge_billing(cash: list, *billing_dicts) -> list:
    combined: dict = {}
    for bd in billing_dicts:
        for lbl, vals in bd.items():
            combined.setdefault(lbl, {"invoiced": 0.0, "collected": 0.0})
            combined[lbl]["invoiced"]  += vals["invoiced"]
            combined[lbl]["collected"] += vals["collected"]

    for m in cash:
        lbl = m["label"]
        if lbl in combined:
            m["invoiced"]  = round(combined[lbl]["invoiced"])
            m["collected"] = round(combined[lbl]["collected"])
    return cash


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Noon Academy ETL → fy_data.json")
    parser.add_argument("--model",           default="Noon financial model 2.0 v42.xlsx")
    parser.add_argument("--cf",              default="Weekly Report CF 2026 Final.xlsx")
    parser.add_argument("--schools-billing", default="Schools_Billing_Updated_May18.xlsx",
                        dest="schools_billing")
    parser.add_argument("--b2b-billing",     default="B2B_Billing_Schedule_Updated_May18.xlsx",
                        dest="b2b_billing")
    parser.add_argument("--mapping",         default="mapping.csv")
    parser.add_argument("--out",             default="fy_data.json")
    args = parser.parse_args()

    model_path   = Path(args.model)
    cf_path      = Path(args.cf)
    schools_path = Path(args.schools_billing)
    b2b_path     = Path(args.b2b_billing)
    out_path     = Path(args.out)

    if not model_path.exists():
        sys.exit(f"ERROR: Model file not found: {model_path}\n"
                 "Make sure you're running from F:\\Noon Dashboard\\ and the filename matches.")

    # ── 1. Financial model ────────────────────────────────────────────────────
    print(f"Loading model: {model_path.name}  (may take ~30 s ...)")
    wb = openpyxl.load_workbook(model_path, data_only=True)

    bu_2425: dict = {}
    bu_2526: dict = {}
    for bu_name, (sheet, rev_labels) in BU_TABS.items():
        if sheet not in wb.sheetnames:
            print(f"  SKIP: sheet '{sheet}' not found — BU will show zeros")
            continue
        ws = wb[sheet]
        print(f"  Extracting {bu_name} from '{sheet}' ...")
        bu_2425[bu_name] = extract_bu(ws, rev_labels, FY2425_COLS)
        bu_2526[bu_name] = extract_bu(ws, rev_labels, FY2526_COLS)

    fy2425_months = build_fy_months(bu_2425, LABELS_2425)
    fy2526_months = build_fy_months(bu_2526, LABELS_2526)

    # ── 2. Cash position ──────────────────────────────────────────────────────
    if cf_path.exists():
        cash = extract_cash(cf_path)
    else:
        print(f"\nWARNING: CF tracker not found ({cf_path}) — cash will be zero.")
        cash = _empty_cash()

    # ── 3. Billing schedules (invoiced / collected) ───────────────────────────
    billing_sources = []

    if schools_path.exists():
        schools_billing = parse_schools_billing(schools_path)
        billing_sources.append(schools_billing)
        print(f"  Schools billing months found: {sorted(schools_billing.keys())}")
    else:
        print(f"\nWARNING: Schools billing file not found ({schools_path}) — skipped.")

    if b2b_path.exists():
        b2b_billing = parse_b2b_billing(b2b_path)
        billing_sources.append(b2b_billing)
    else:
        print(f"\nWARNING: B2B billing file not found ({b2b_path}) — skipped.")

    if billing_sources:
        cash = merge_billing(cash, *billing_sources)
        print("\nBilling merged into cash array:")
        for m in cash:
            if m["invoiced"] or m["collected"]:
                print(f"  {m['label']:8s}  invoiced=${m['invoiced']:>9,.0f}  "
                      f"collected=${m['collected']:>9,.0f}")

    # ── 4. Write output ───────────────────────────────────────────────────────
    billing_note = (
        "Invoiced/collected from Schools_Billing + B2B_Billing (SAR÷3.75→USD). "
        "Confirm counterparty→BU mapping with Henry."
    ) if billing_sources else (
        "Invoiced/collected not available — billing files not found."
    )

    output = {
        "meta": {
            "generated_at": date.today().isoformat(),
            "source": f"Extracted from {model_path.name}",
            "billing_note": billing_note,
        },
        "buses":  list(BU_TABS.keys()),
        "fy2425": fy2425_months,
        "fy2526": fy2526_months,
        "cash":   cash,
    }

    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)

    size_kb = out_path.stat().st_size / 1024
    print(f"\nWrote {out_path}  ({size_kb:.0f} KB)")
    print("Launch dashboard:  python -m streamlit run app.py")


if __name__ == "__main__":
    main()
