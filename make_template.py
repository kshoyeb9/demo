"""
make_template.py — Generates the Noon dashboard input workbook.

Two tabs: 'Instructions' and 'Dashboard Input'.

Structure (v6):
  The twelve-month grids are the single source of truth. Every month and
  year-to-date figure is a formula that rolls up from them, cut off at the
  active month set in Section 1 — so closing a month means changing ONE cell,
  not re-pointing any formulas.

  etl_pl.py deliberately does not read those formula cells. It re-derives the
  same figures from the same grids, so the dashboard is correct whether or not
  Excel has recalculated the file.

Usage:
    python make_template.py [--out noon_dashboard_input.xlsx]
"""

import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

ap = argparse.ArgumentParser()
ap.add_argument("--out", default="noon_dashboard_input.xlsx")
args = ap.parse_args()

SHEET = "Dashboard Input"

# ── Style tokens ──────────────────────────────────────────────────────────────
FONT      = "Verdana"
C_HEADER  = "11203A"
C_SUBHDR  = "D9E1F2"
C_INPUT   = "FFF3CD"
C_CALC    = "F2F2F2"
C_TOTAL   = "BDD7EE"
C_WHITE   = "FFFFFF"
BLUE_TXT  = "0000FF"
BLACK_TXT = "000000"

FMT_M   = '$#,##0.00;($#,##0.00);-'
FMT_MS  = '+$#,##0.00;-$#,##0.00;-'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0.0'

MONTHS   = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
YR       = "26"
ACTIVE   = 6          # June — the last closed month in this build
NM       = len(MONTHS)
C0, C1   = 2, 1 + NM  # month columns B..M
LAST_COL = C1 + 1

def F(bold=False, color=BLACK_TXT, size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
def fill(h): return PatternFill("solid", fgColor=h)
def A(h="left", wrap=False, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Side(style="thin", color="BFBFBF")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook(); wb.remove(wb.active)

def bar(ws, r, text, width=LAST_COL):
    c = ws.cell(row=r, column=1, value=text)
    c.font = F(bold=True, color=C_WHITE); c.alignment = A()
    for cc in range(1, width+1): ws.cell(row=r, column=cc).fill = fill(C_HEADER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    ws.row_dimensions[r].height = 20

def note(ws, r, text, width=LAST_COL):
    c = ws.cell(row=r, column=1, value=text)
    c.font = F(size=10, italic=True, color="595959"); c.alignment = A()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)

def headers(ws, r, labels, start=1, merges=None):
    merges = merges or {}; col = start
    for i, t in enumerate(labels):
        ws.cell(row=r, column=col, value=t)
        span = merges.get(i, 1)
        for cc in range(col, col+span):
            x = ws.cell(row=r, column=cc)
            x.font = F(bold=True); x.fill = fill(C_SUBHDR); x.border = BOX
            x.alignment = A("center", wrap=True)
        if span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+span-1)
        col += span
    ws.row_dimensions[r].height = 26

def put(ws, r, c, v, *, kind="input", fmt=None, bold=False, wrap=False, halign=None, span=1):
    bg = {"input": C_INPUT, "calc": C_CALC, "total": C_TOTAL}.get(kind, C_WHITE)
    for cc in range(c, c+span):
        x = ws.cell(row=r, column=cc); x.fill = fill(bg); x.border = BOX
    cell = ws.cell(row=r, column=c, value=v)
    isnum = isinstance(v, (int, float))
    cell.font = F(bold=bold or kind == "total",
                  color=BLUE_TXT if kind == "input" and isnum else BLACK_TXT)
    if fmt: cell.number_format = fmt
    cell.alignment = A(halign or ("right" if isnum else "left"), wrap=wrap)
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    return cell


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — INSTRUCTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Instructions")
for col, w in {"A": 4, "B": 30, "C": 92, "D": 30}.items():
    ws.column_dimensions[col].width = w

ws["B2"] = "Noon Academy — Dashboard Input Workbook"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color=C_HEADER)
ws["B3"] = "Fill in the 'Dashboard Input' tab, then regenerate the dashboard."
ws["B3"].font = F(italic=True, color="595959")

def ibar(r, text):
    c = ws.cell(row=r, column=2, value=text); c.font = F(bold=True, color=C_WHITE)
    for cc in (2,3,4): ws.cell(row=r, column=cc).fill = fill(C_HEADER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 20

r = 5
ibar(r, "  CLOSING A MONTH"); r += 2
for title, body in [
    ("1.  Enter the month",
     "On the 'Dashboard Input' tab, type the month's figures into the twelve-month grids: "
     "Section 5 (revenue and cost actuals), Section 6 (working capital), Sections 7 and 8 "
     "(receivables and payables), and Section 9 (cash flow). These grids are the only place "
     "actuals are entered."),
    ("2.  Move the active month",
     "In Section 1, set 'Active month number' to the month you just closed — 6 for June, "
     "7 for July, and so on. Every month and year-to-date figure in the workbook rolls up to "
     "that month and stops. Nothing else needs re-pointing."),
    ("3.  Update the labels",
     "Still in Section 1, update 'Period label' and 'YTD label' to match. These are text only; "
     "they appear in the dashboard headers and are not used in any calculation."),
    ("4.  Regenerate",
     "Save and close, then run:      python etl_pl.py --file noon_dashboard_input.xlsx"),
]:
    ws.cell(row=r, column=2, value=title).font = F(bold=True, size=11)
    c = ws.cell(row=r, column=3, value=body); c.font = F(); c.alignment = A(wrap=True, v="top")
    ws.row_dimensions[r].height = 52; r += 1

r += 1
ibar(r, "  WHERE NUMBERS COME FROM"); r += 1
for i, t in enumerate(["", "Section", "What it holds", "Derived from"]):
    if not i: continue
    c = ws.cell(row=r, column=1+i, value=t)
    c.font = F(bold=True); c.fill = fill(C_SUBHDR); c.border = BOX; c.alignment = A("center")
r += 1
for name, holds, src in [
    ("1 · Setup",            "Period labels and the active month number.", "You"),
    ("2 · Revenue by BU",    "Month and YTD, actual vs budget.",           "Rolls up from Section 5"),
    ("3 · Costs",            "Month and YTD, actual vs budget.",           "Rolls up from Section 5"),
    ("4 · P&L summary",      "Gross profit, contribution, EBITDA, margins.","Sections 2 and 3"),
    ("5 · Monthly grids",    "Twelve months of revenue and cost, actual and budget.", "You"),
    ("6 · Working capital",  "Twelve months of closing balances; YTD movement.", "You (monthly), rest rolls up"),
    ("7 · Receivables",      "Twelve months of invoiced and collected; aging; by contract.", "You (monthly), rest rolls up"),
    ("8 · Payables",         "Twelve months of purchases and payments; aging; by vendor.", "You (monthly), rest rolls up"),
    ("9 · Cash flow",        "Twelve months of inflows and outflows.",     "You"),
    ("10 · Cash position",   "KPI tiles, both bridges, monthly balance, runway.", "All rolls up from Section 9"),
    ("11 · Key updates",     "Narrative commentary.",                      "You"),
]:
    put(ws, r, 2, name, kind="plain", bold=True)
    put(ws, r, 3, holds, kind="plain", wrap=True)
    put(ws, r, 4, src,   kind="plain", wrap=True)
    ws.row_dimensions[r].height = 28; r += 1

r += 1
ibar(r, "  CELL COLOURS"); r += 1
for hexc, nm, mean in [
    (C_INPUT, "Yellow", "An input. Type here."),
    (C_CALC,  "Grey",   "A formula that rolls up from the grids. Do not overwrite — you would break the roll-up."),
    (C_TOTAL, "Blue",   "A total. Also a formula."),
    (C_WHITE, "White",  "A label."),
]:
    c = ws.cell(row=r, column=2, value=nm); c.fill = fill(hexc); c.font = F(bold=True); c.border = BOX
    put(ws, r, 3, mean, kind="plain", wrap=True); r += 1

r += 1
ibar(r, "  CONVENTIONS"); r += 1
for k, v in [
    ("Units",        "USD millions. Enter 2.5 for $2.5M — not 2500000."),
    ("Currency",     "Always enter USD, whatever the source ledger. The dashboard has a USD/SAR "
                     "toggle that converts on display at 1 USD = 3.75 SAR; entering SAR here would double-convert."),
    ("Costs",        "Enter costs and outflows as POSITIVE numbers. Signs are applied for you."),
    ("Working capital", "Payables and deferred revenue are NEGATIVE — they are liabilities."),
    ("Future months","Leave actuals blank beyond the active month. Budget is filled for all twelve."),
    ("Percentages",  "Enter as a percentage-formatted number (96.0%), not 0.96 or 96."),
    ("Blank rows",   "A blank row ends a table. Never leave a gap mid-table."),
    ("Headers",      "Do not rename or delete a header row — tables are found by their header text."),
]:
    put(ws, r, 2, k, kind="plain", bold=True)
    put(ws, r, 3, v, kind="plain", wrap=True)
    ws.row_dimensions[r].height = 30; r += 1

ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — DASHBOARD INPUT
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet(SHEET)
ws.column_dimensions["A"].width = 44
for i in range(C0, LAST_COL+2):
    ws.column_dimensions[CL(i)].width = 12.5

r = 1

# ── 1 · SETUP ─────────────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 1 — SETUP"); r += 1
note(ws, r, "Set the active month first: everything in the workbook rolls up to it and stops."); r += 1
headers(ws, r, ["Setting", "Value", "Notes"], merges={2: 4}); r += 1
SETTINGS = [
    ("Period label (month)",    "June 2026",    "Text only — the latest closed month"),
    ("YTD label",               "Jan–Jun 2026", "Text only — range covered year-to-date"),
    ("Reporting year",          "2026",         "Calendar year of the current period"),
    ("Fiscal year label",       "FY2025/26",    "Shown next to quarter labels"),
    ("Fiscal year start month", 7,              "1 = January … 7 = July"),
    ("Currency label",          "USD M",        "Displayed under chart titles"),
    ("Active month number",     ACTIVE,         "1 = Jan … 12 = Dec. THE control: every roll-up stops here."),
]
setting_row = {}
for k, v, n in SETTINGS:
    setting_row[k] = r
    put(ws, r, 1, k, kind="plain", bold=True)
    put(ws, r, 2, v, kind="input")
    put(ws, r, 3, n, kind="plain", span=4)
    r += 1
ACT = f"$B${setting_row['Active month number']}"     # absolute ref used by every roll-up
r += 1

note(ws, r, "The twelve months of the reporting year, in order."); r += 1
headers(ws, r, ["Setup Month", "Short label"]); r += 1
for m in MONTHS:
    put(ws, r, 1, f"{m}-{YR}", kind="input")
    put(ws, r, 2, m, kind="input")
    r += 1
r += 1

# ── helpers for the monthly grids ─────────────────────────────────────────────
def month_row_refs(row):
    """(whole-row range, roll-up-to-active range) for one twelve-month row."""
    full = f"${CL(C0)}{row}:${CL(C1)}{row}"
    return full, f"${CL(C0)}{row}:INDEX({full},{ACT})"

def val_at_active(row):
    full, _ = month_row_refs(row)
    return f"=INDEX({full},{ACT})"

def sum_to_active(row):
    _, upto = month_row_refs(row)
    return f"=SUM({upto})"

def col_at_active(col_letter, first, last):
    """Value in `col_letter` on the active month's row, for month-per-row tables."""
    return f"=INDEX(${col_letter}${first}:${col_letter}${last},{ACT})"

def col_sum_to_active(col_letter, first, last):
    return f"=SUM(${col_letter}${first}:INDEX(${col_letter}${first}:${col_letter}${last},{ACT}))"


# ── 5 · MONTHLY GRIDS (built first — sections 2 and 3 point at them) ──────────
REVENUE = ["Tracks", "B2B", "Govt Schools — Legacy", "Govt Schools — New", "Out of School"]
COSTS   = ["Direct costs", "Marketing", "BU salaries", "Noon HQ", "Other operating expenses"]

REV_ACT = {  # Jan–Jun actual; Jul–Dec blank
    "Tracks":                [0.876,0.904,0.960,0.932,0.988,0.990],
    "B2B":                   [0.566,0.584,0.620,0.602,0.638,0.630],
    "Govt Schools — Legacy":  [0.461,0.475,0.505,0.490,0.519,0.520],
    "Govt Schools — New":     [0.216,0.223,0.237,0.230,0.244,0.260],
    "Out of School":         [0.109,0.113,0.120,0.116,0.123,0.100],
}
REV_BGT = {  # all twelve months
    "Tracks":                [0.959,0.989,1.051,1.020,1.081,1.070,1.085,1.100,1.120,1.140,1.160,1.180],
    "B2B":                   [0.581,0.600,0.637,0.618,0.655,0.630,0.645,0.660,0.675,0.690,0.705,0.720],
    "Govt Schools — Legacy":  [0.496,0.512,0.544,0.528,0.560,0.560,0.565,0.570,0.575,0.580,0.585,0.590],
    "Govt Schools — New":     [0.239,0.246,0.262,0.254,0.269,0.290,0.300,0.310,0.320,0.330,0.340,0.350],
    "Out of School":         [0.094,0.097,0.103,0.100,0.106,0.100,0.105,0.110,0.115,0.120,0.125,0.130],
}
COST_ACT = {
    "Direct costs":             [0.784,0.809,0.859,0.834,0.884,0.890],
    "Marketing":                [0.107,0.111,0.117,0.114,0.121,0.100],
    "BU salaries":              [0.600,0.619,0.657,0.638,0.676,0.680],
    "Noon HQ":                  [0.182,0.188,0.200,0.194,0.206,0.210],
    "Other operating expenses": [0.216,0.223,0.237,0.230,0.244,0.260],
}
COST_BGT = {
    "Direct costs":             [0.820,0.846,0.898,0.872,0.924,0.920,0.930,0.940,0.950,0.960,0.970,0.980],
    "Marketing":                [0.130,0.134,0.142,0.138,0.146,0.130,0.135,0.140,0.145,0.150,0.155,0.160],
    "BU salaries":              [0.609,0.629,0.667,0.648,0.687,0.700,0.710,0.720,0.730,0.740,0.750,0.760],
    "Noon HQ":                  [0.188,0.194,0.206,0.200,0.212,0.200,0.205,0.210,0.215,0.220,0.225,0.230],
    "Other operating expenses": [0.220,0.227,0.241,0.234,0.248,0.240,0.245,0.250,0.255,0.260,0.265,0.270],
}

monthly_start = r
bar(ws, r, "  SECTION 5 — MONTHLY ACTUAL & BUDGET  (USD M)"); r += 1
note(ws, r, "The source of truth. Actuals stop at the active month; budget covers all twelve."); r += 1

grid_row = {}   # (block, line name) -> row
for block, data, is_actual in [
    ("Revenue — Actual", REV_ACT,  True),
    ("Revenue — Budget", REV_BGT,  False),
    ("Costs — Actual",   COST_ACT, True),
    ("Costs — Budget",   COST_BGT, False),
]:
    headers(ws, r, [block] + MONTHS); r += 1
    first = r
    for name, vals in data.items():
        put(ws, r, 1, name, kind="plain")
        for i in range(NM):
            v = vals[i] if i < len(vals) else None
            put(ws, r, C0+i, v, kind="input", fmt=FMT_M)
        grid_row[(block, name)] = r
        r += 1
    last = r-1
    put(ws, r, 1, "Total", kind="total")
    for i in range(NM):
        L = CL(C0+i)
        put(ws, r, C0+i, f"=SUM({L}{first}:{L}{last})", kind="total", fmt=FMT_M)
    grid_row[(block, "Total")] = r
    r += 2

# ── 2 & 3 · REVENUE AND COSTS (all four columns roll up) ─────────────────────
AVB = ["Actual (Month)", "Budget (Month)", "Actual YTD", "Budget YTD"]

def avb_block(r, title, hdr0, names, act_block, bgt_block, total_label):
    bar(ws, r, title, width=5); r += 1
    note(ws, r, "Every figure rolls up from Section 5, cut off at the active month.", width=5); r += 1
    headers(ws, r, [hdr0] + AVB); r += 1
    first = r
    rows = {}
    for n in names:
        ar_, br_ = grid_row[(act_block, n)], grid_row[(bgt_block, n)]
        put(ws, r, 1, n, kind="plain")
        put(ws, r, 2, val_at_active(ar_),  kind="calc", fmt=FMT_M)
        put(ws, r, 3, val_at_active(br_),  kind="calc", fmt=FMT_M)
        put(ws, r, 4, sum_to_active(ar_),  kind="calc", fmt=FMT_M)
        put(ws, r, 5, sum_to_active(br_),  kind="calc", fmt=FMT_M)
        rows[n] = r; r += 1
    last = r-1
    put(ws, r, 1, total_label, kind="total")
    for c in range(2, 6):
        L = CL(c); put(ws, r, c, f"=SUM({L}{first}:{L}{last})", kind="total", fmt=FMT_M)
    return r+2, rows, r

sec2 = r
r, REV_ROW, rev_total_row = avb_block(
    r, "  SECTION 2 — REVENUE BY BUSINESS UNIT  (USD M)", "Business Unit",
    REVENUE, "Revenue — Actual", "Revenue — Budget", "Total Revenue")
r, COST_ROW, cost_total_row = avb_block(
    r, "  SECTION 3 — COSTS BY CATEGORY  (USD M)", "Cost Category",
    COSTS, "Costs — Actual", "Costs — Budget", "Total costs")

# ── 4 · P&L SUMMARY ───────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 4 — P&L SUMMARY  (USD M)", width=7); r += 1
note(ws, r, "Pulls from Sections 2 and 3. Nothing to enter.", width=7); r += 1
headers(ws, r, ["P&L Line"] + AVB + ["Type", "Is a cost?"]); r += 1

def src(row): return tuple(f"={CL(c)}{row}" for c in range(2, 6))
PL = [
    ("Revenue",                  *src(rev_total_row),              "Value", "No"),
    ("Direct costs",             *src(COST_ROW["Direct costs"]),   "Value", "Yes"),
    ("Gross profit",             None,None,None,None,              "Value", "No"),
    ("Marketing expense",        *src(COST_ROW["Marketing"]),      "Value", "Yes"),
    ("Contribution profit",      None,None,None,None,              "Value", "No"),
    ("BU salaries",              *src(COST_ROW["BU salaries"]),    "Value", "Yes"),
    ("Noon HQ",                  *src(COST_ROW["Noon HQ"]),        "Value", "Yes"),
    ("Other operating expenses", *src(COST_ROW["Other operating expenses"]), "Value", "Yes"),
    ("EBITDA",                   None,None,None,None,              "Value", "No"),
]
pl_row = {}
for name, b, c_, d, e, kind, cost in PL:
    pl_row[name] = r
    sub = name in ("Gross profit", "Contribution profit", "EBITDA")
    put(ws, r, 1, name, kind="plain", bold=sub)
    if b is not None:
        for col, f_ in zip(range(2, 6), (b, c_, d, e)):
            put(ws, r, col, f_, kind="calc", fmt=FMT_M)
    put(ws, r, 6, kind, kind="plain", halign="center")
    put(ws, r, 7, cost, kind="plain", halign="center")
    r += 1
gp, cp, eb = pl_row["Gross profit"], pl_row["Contribution profit"], pl_row["EBITDA"]
rv, dc, mk = pl_row["Revenue"], pl_row["Direct costs"], pl_row["Marketing expense"]
sl, hq, ot = pl_row["BU salaries"], pl_row["Noon HQ"], pl_row["Other operating expenses"]
for c in range(2, 6):
    L = CL(c)
    put(ws, gp, c, f"={L}{rv}-{L}{dc}",                 kind="total", fmt=FMT_M)
    put(ws, cp, c, f"={L}{gp}-{L}{mk}",                 kind="total", fmt=FMT_M)
    put(ws, eb, c, f"={L}{cp}-{L}{sl}-{L}{hq}-{L}{ot}", kind="total", fmt=FMT_M)
r += 1
headers(ws, r, ["Margin"] + AVB); r += 1
for label, num in (("Gross profit margin", gp), ("Contribution margin", cp), ("EBITDA margin", eb)):
    put(ws, r, 1, label, kind="plain")
    for c in range(2, 6):
        L = CL(c)
        put(ws, r, c, f'=IF({L}{rv}=0,"",{L}{num}/{L}{rv})', kind="calc", fmt=FMT_PCT)
    r += 1
r += 1

# ── 6 · WORKING CAPITAL ───────────────────────────────────────────────────────
bar(ws, r, "  SECTION 6 — WORKING CAPITAL  (USD M)", width=8); r += 1
note(ws, r, "Closing balances. Payables and deferred revenue are NEGATIVE. Blank beyond the active month.", width=8); r += 1
headers(ws, r, ["WC Month","Receivables","Payables","Deferred revenue","Other WC",
                "Net WC","Movement","Forecast?"]); r += 1
WC = [(5.9,-2.9,-2.4,0.5),(6.1,-2.8,-2.3,0.5),(6.3,-2.7,-2.2,0.6),
      (6.4,-2.9,-2.2,0.5),(6.6,-2.8,-2.1,0.6),(6.7,-2.7,-2.1,0.6)]
wc_first = r
for i, m in enumerate(MONTHS):
    vals = WC[i] if i < len(WC) else (None,)*4
    put(ws, r, 1, f"{m}-{YR}", kind="input")
    for j, v in enumerate(vals): put(ws, r, C0+j, v, kind="input", fmt=FMT_M)
    put(ws, r, 6, f"=IF(COUNT(B{r}:E{r})=0,\"\",SUM(B{r}:E{r}))", kind="calc", fmt=FMT_M)
    put(ws, r, 7, 0 if i == 0 else f'=IF(OR(F{r}="",F{r-1}=""),"",F{r}-F{r-1})', kind="calc", fmt=FMT_MS)
    put(ws, r, 8, f'=IF({i+1}>{ACT},"Yes","No")', kind="calc", halign="center")
    r += 1
wc_last = r-1
r += 1

note(ws, r, "Opening is 1 January; period end rolls up to the active month.", width=5); r += 1
headers(ws, r, ["WC Item","At 1 Jan","At period end","Movement","Cash impact"]); r += 1
wy_first = r
for i, item in enumerate(["Accounts receivable","Accounts payable","Deferred revenue","Other working capital items"]):
    put(ws, r, 1, item, kind="plain")
    put(ws, r, 2, f"={CL(C0+i)}{wc_first}", kind="calc", fmt=FMT_M)
    put(ws, r, 3, col_at_active(CL(C0+i), wc_first, wc_last), kind="calc", fmt=FMT_M)
    put(ws, r, 4, f"=C{r}-B{r}",    kind="calc", fmt=FMT_MS)
    put(ws, r, 5, f"=-(C{r}-B{r})", kind="calc", fmt=FMT_MS)
    r += 1
wy_last = r-1
put(ws, r, 1, "Net working capital", kind="total")
for c in range(2, 6):
    L = CL(c)
    put(ws, r, c, f"=SUM({L}{wy_first}:{L}{wy_last})", kind="total",
        fmt=FMT_M if c in (2,3) else FMT_MS)
r += 2

# ── 7 · AR ────────────────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 7 — ACCOUNTS RECEIVABLE  (USD M)", width=6); r += 1
note(ws, r, "Enter invoiced and collected. Opening chains from the prior month; closing and rate calculate.", width=6); r += 1
headers(ws, r, ["AR Month","Opening AR","Invoiced","Collected","Closing AR","Collection rate"]); r += 1
AR = [(2.30,2.10),(2.25,2.05),(2.55,2.35),(2.40,2.30),(2.35,2.15),(2.50,2.40)]
ar_first = r
for i, m in enumerate(MONTHS):
    inv, coll = AR[i] if i < len(AR) else (None, None)
    put(ws, r, 1, f"{m}-{YR}", kind="input")
    put(ws, r, 2, 5.70 if i == 0 else f'=IF(E{r-1}="","",E{r-1})',
        kind="input" if i == 0 else "calc", fmt=FMT_M)
    put(ws, r, 3, inv,  kind="input", fmt=FMT_M)
    put(ws, r, 4, coll, kind="input", fmt=FMT_M)
    put(ws, r, 5, f'=IF(COUNT(C{r}:D{r})=0,"",B{r}+C{r}-D{r})', kind="calc", fmt=FMT_M)
    put(ws, r, 6, f'=IF(N(C{r})=0,"",D{r}/C{r})', kind="calc", fmt=FMT_PCT)
    r += 1
ar_last = r-1
put(ws, r, 1, "YTD total", kind="total")
put(ws, r, 2, "", kind="total")
put(ws, r, 3, col_sum_to_active("C", ar_first, ar_last), kind="total", fmt=FMT_M)
put(ws, r, 4, col_sum_to_active("D", ar_first, ar_last), kind="total", fmt=FMT_M)
put(ws, r, 5, col_at_active("E", ar_first, ar_last),     kind="total", fmt=FMT_M)
put(ws, r, 6, f'=IF(N(C{r})=0,"",D{r}/C{r})',            kind="total", fmt=FMT_PCT)
r += 2

def share_table(r, note_text, hdr, rows_data, total_label, name_kind="plain"):
    note(ws, r, note_text, width=3); r += 1
    headers(ws, r, [hdr, "Amount", "Share"]); r += 1
    first = r
    for label, amt in rows_data:
        put(ws, r, 1, label, kind=name_kind); put(ws, r, 2, amt, kind="input", fmt=FMT_M); r += 1
    last, tot = r-1, r
    for rr in range(first, last+1):
        put(ws, rr, 3, f'=IF($B${tot}=0,"",B{rr}/$B${tot})', kind="calc", fmt=FMT_PCT)
    put(ws, tot, 1, total_label, kind="total")
    put(ws, tot, 2, f"=SUM(B{first}:B{last})", kind="total", fmt=FMT_M)
    put(ws, tot, 3, 1.0, kind="total", fmt=FMT_PCT)
    return tot+2

r = share_table(r, "Balance at the active month, split by age.", "AR Aging Bucket",
                [("Current — due next 30 days",2.8),("31–60 days",2.0),("60+ days / overdue",1.9)],
                "Total AR")
r = share_table(r, "Largest receivable balances by contract.", "Contract",
                [("MCIT",2.10),("Takaful",1.50),("Taalum",1.20),
                 ("Ensan",0.80),("Tracks",0.60),("Other contracts",0.50)],
                "Total", name_kind="input")

# ── 8 · AP ────────────────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 8 — ACCOUNTS PAYABLE  (USD M)", width=6); r += 1
note(ws, r, "Enter purchases, payments and DPO. Opening chains from the prior month.", width=6); r += 1
headers(ws, r, ["AP Month","Opening AP","Purchases","Payments","Closing AP","DPO (days)"]); r += 1
AP = [(2.00,2.05,43.5),(1.95,2.05,43.1),(2.00,2.10,40.5),(2.10,1.90,41.4),(1.95,2.05,43.1),(2.05,2.15,39.5)]
ap_first = r
for i, m in enumerate(MONTHS):
    pur, pay, dpo = AP[i] if i < len(AP) else (None, None, None)
    put(ws, r, 1, f"{m}-{YR}", kind="input")
    put(ws, r, 2, 2.95 if i == 0 else f'=IF(E{r-1}="","",E{r-1})',
        kind="input" if i == 0 else "calc", fmt=FMT_M)
    put(ws, r, 3, pur, kind="input", fmt=FMT_M)
    put(ws, r, 4, pay, kind="input", fmt=FMT_M)
    put(ws, r, 5, f'=IF(COUNT(C{r}:D{r})=0,"",B{r}+C{r}-D{r})', kind="calc", fmt=FMT_M)
    put(ws, r, 6, dpo, kind="input", fmt=FMT_NUM)
    r += 1
ap_last = r-1
put(ws, r, 1, "YTD total", kind="total")
put(ws, r, 2, "", kind="total")
put(ws, r, 3, col_sum_to_active("C", ap_first, ap_last), kind="total", fmt=FMT_M)
put(ws, r, 4, col_sum_to_active("D", ap_first, ap_last), kind="total", fmt=FMT_M)
put(ws, r, 5, col_at_active("E", ap_first, ap_last),     kind="total", fmt=FMT_M)
put(ws, r, 6, f'=IF({ACT}=0,"",AVERAGE(F{ap_first}:INDEX(F{ap_first}:F{ap_last},{ACT})))',
    kind="total", fmt=FMT_NUM)
r += 2

r = share_table(r, "Balance at the active month, split by age.", "AP Aging Bucket",
                [("Current — due next 30 days",1.5),("31–60 days",0.8),("60+ days / overdue",0.4)],
                "Total AP")
r = share_table(r, "Largest payable balances by vendor.", "Vendor",
                [("AWS / cloud infrastructure",0.62),("Content production partners",0.48),
                 ("Facilities & office leases",0.37),("Marketing agencies",0.29),
                 ("Professional services",0.21),("Other vendors",0.18)],
                "Total", name_kind="input")

# ── 9 · MONTHLY CASH FLOW (drives everything in section 10) ──────────────────
bar(ws, r, "  SECTION 9 — MONTHLY CASH FLOW  (USD M)"); r += 1
note(ws, r, "Enter every amount POSITIVE. Opening cash is entered once, for January; each later month chains."); r += 1
note(ws, r, "Section 10 — tiles, both bridges, the balance chart and runway — is calculated entirely from this grid."); r += 1
headers(ws, r, ["Cash Flow Line"] + MONTHS); r += 1

CF_IN  = [("Collections",   [2.10,2.05,2.35,2.30,2.15,2.40]),
          ("Other inflows", [0.08,0.10,0.12,0.09,0.10,0.10])]
CF_OUT = [("Operating costs",[1.95,2.00,2.05,2.02,2.00,2.08]),
          ("Capex",          [0.14,0.15,0.16,0.15,0.15,0.16]),
          ("Debt service",   [0.15,0.15,0.15,0.15,0.15,0.16]),
          ("Other outflows", [0.12,0.10,0.15,0.12,0.13,0.42])]
cf_row = {}

put(ws, r, 1, "Opening cash", kind="plain", bold=True)
for i in range(NM):
    put(ws, r, C0+i, 4.80 if i == 0 else f'=IF({CL(C0+i-1)}{r+len(CF_IN)+len(CF_OUT)+2}="","",{CL(C0+i-1)}{r+len(CF_IN)+len(CF_OUT)+2})',
        kind="input" if i == 0 else "calc", fmt=FMT_M)
cf_row["Opening cash"] = r; r += 1

for name, vals in CF_IN + CF_OUT:
    put(ws, r, 1, name, kind="plain")
    for i in range(NM):
        put(ws, r, C0+i, vals[i] if i < len(vals) else None, kind="input", fmt=FMT_M)
    cf_row[name] = r; r += 1

in_rows  = [cf_row[n] for n, _ in CF_IN]
out_rows = [cf_row[n] for n, _ in CF_OUT]
put(ws, r, 1, "Net movement", kind="total")
for i in range(NM):
    L = CL(C0+i)
    plus  = "+".join(f"N({L}{x})" for x in in_rows)
    minus = "+".join(f"N({L}{x})" for x in out_rows)
    put(ws, r, C0+i, f'=IF(COUNT({L}{in_rows[0]}:{L}{out_rows[-1]})=0,"",({plus})-({minus}))',
        kind="total", fmt=FMT_MS)
cf_row["Net movement"] = r; r += 1

put(ws, r, 1, "Closing cash", kind="total")
for i in range(NM):
    L = CL(C0+i)
    put(ws, r, C0+i, f'=IF({L}{cf_row["Net movement"]}="","",{L}{cf_row["Opening cash"]}+{L}{cf_row["Net movement"]})',
        kind="total", fmt=FMT_M)
cf_row["Closing cash"] = r; r += 2

CFV = lambda name: val_at_active(cf_row[name])          # value at active month
CFS = lambda name: sum_to_active(cf_row[name])          # sum to active month

# ── 10 · CASH POSITION ────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 10 — CASH POSITION  (USD M)", width=9); r += 1
note(ws, r, "Every figure below is calculated from Section 9. Only the tile notes are text you write.", width=9); r += 1
headers(ws, r, ["Cash Tile","Month value","Month note","YTD value","YTD note"],
        merges={2: 3, 4: 3}); r += 1
close_act  = CFV("Closing cash")
open_jan   = f"={CL(C0)}{cf_row['Opening cash']}"
nwc_act    = col_at_active("F", wc_first, wc_last)
ar_act     = col_at_active("E", ar_first, ar_last)
for name, mv, mn, yv, yn in [
    ("Cash balance",        close_act, "months of runway at the current burn",
                            close_act, "against the 1 January opening balance"),
    ("Collections",         CFV("Collections"), "collected in the month",
                            CFS("Collections"), "collected year-to-date"),
    ("Net working capital", nwc_act, "net position at the active month",
                            f"={nwc_act[1:]}-{CL(C0)}{wc_first}", "movement since 1 January"),
    ("Accounts receivable", ar_act, "closing receivables",
                            ar_act, "closing receivables"),
]:
    put(ws, r, 1, name, kind="plain", bold=True)
    put(ws, r, 2, mv, kind="calc", fmt=FMT_M)
    put(ws, r, 3, mn, kind="input", span=3, wrap=True)
    put(ws, r, 6, yv, kind="calc", fmt=FMT_M)
    put(ws, r, 7, yn, kind="input", span=3, wrap=True)
    r += 1
r += 1

def bridge(r, title, hdr, per_month):
    """per_month True -> the active month's column; False -> Jan..active."""
    note(ws, r, "Calculated from Section 9. Type is fixed; amounts roll up.", width=4); r += 1
    headers(ws, r, [hdr, "Type", "Amount", "Running balance"]); r += 1
    first = r
    steps = [("Opening cash", "Opening", f"=IF({ACT}<=1,{CL(C0)}{cf_row['Opening cash']},INDEX({month_row_refs(cf_row['Opening cash'])[0]},{ACT}))"
                                          if per_month else f"={CL(C0)}{cf_row['Opening cash']}")]
    for n, _ in CF_IN:  steps.append((n, "Inflow",  CFV(n) if per_month else CFS(n)))
    for n, _ in CF_OUT: steps.append((n, "Outflow", CFV(n) if per_month else CFS(n)))
    steps.append(("Closing cash", "Closing", CFV("Closing cash")))
    for i, (step, kind_, amt) in enumerate(steps):
        put(ws, r, 1, step,  kind="plain")
        put(ws, r, 2, kind_, kind="plain", halign="center")
        put(ws, r, 3, amt,   kind="calc", fmt=FMT_M)
        put(ws, r, 4, f"=C{r}" if i == 0 else
            f'=IF(B{r}="Closing",D{r-1},IF(B{r}="Inflow",D{r-1}+C{r},D{r-1}-C{r}))',
            kind="calc", fmt=FMT_M)
        r += 1
    return r+1

r = bridge(r, "", "Bridge Step (Month)", True)
r = bridge(r, "", "Bridge Step (YTD)",   False)

note(ws, r, "Drives the cash balance chart. Prior-year months are typed; this year's roll up from Section 9."); r += 1
headers(ws, r, ["Cash Month","Closing cash","Forecast?","Illustrative?"]); r += 1
for m, v in [("Jul-25",5.35),("Aug-25",5.22),("Sep-25",5.10),
             ("Oct-25",5.02),("Nov-25",4.91),("Dec-25",4.80)]:
    put(ws, r, 1, m, kind="input"); put(ws, r, 2, v, kind="input", fmt=FMT_M)
    put(ws, r, 3, "No", kind="input", halign="center")
    put(ws, r, 4, "Yes" if m != "Dec-25" else "No", kind="input", halign="center")
    r += 1
for i, m in enumerate(MONTHS):
    put(ws, r, 1, f"{m}-{YR}", kind="input")
    put(ws, r, 2, f"={CL(C0+i)}{cf_row['Closing cash']}", kind="calc", fmt=FMT_M)
    put(ws, r, 3, f'=IF({i+1}>{ACT},"Yes","No")', kind="calc", halign="center")
    put(ws, r, 4, "No", kind="input", halign="center")
    r += 1
r += 1

note(ws, r, "All four roll up from Section 9. Burn is negative when cash is consumed.", width=5); r += 1
headers(ws, r, ["Runway Metric","Value","Notes"], merges={2: 3}); r += 1
nm_row = cf_row["Net movement"]
nm_full = month_row_refs(nm_row)[0]
burn3 = (f'=IF({ACT}<3,INDEX({nm_full},{ACT}),'
         f'AVERAGE(INDEX({nm_full},MAX(1,{ACT}-2)):INDEX({nm_full},{ACT})))')
for metric, formula, nt in [
    ("Monthly cash burn",             CFV("Net movement"), "Net movement in the active month"),
    ("Trailing 3-month average burn", burn3,               "Average net movement over the last three months"),
    ("YTD net cash movement",         CFS("Net movement"), "Sum of net movement, January to the active month"),
    ("Cash runway (months)",          None,                "Closing cash ÷ the 3-month average burn"),
]:
    put(ws, r, 1, metric, kind="plain")
    put(ws, r, 2, formula, kind="calc", fmt=FMT_M if "months" not in metric else FMT_NUM)
    put(ws, r, 3, nt, kind="plain", span=3)
    r += 1
runway_first = r-4
put(ws, runway_first+3, 2,
    f'=IF(B{runway_first+1}>=0,"",{close_act[1:]}/-B{runway_first+1})',
    kind="calc", fmt=FMT_NUM)
r += 1

# ── 11 · KEY UPDATES ──────────────────────────────────────────────────────────
bar(ws, r, "  SECTION 11 — KEY NARRATIVE UPDATES"); r += 1
note(ws, r, "One row per commentary point, in the order shown on the dashboard."); r += 1
headers(ws, r, ["Update #","Topic","Commentary"], merges={2: 11}); r += 1
for i, (topic, text) in enumerate([
    ("Revenue","June revenue of $2.50M landed at 94.3% of budget, missing plan in five of the first six months of the year. YTD revenue of $14.35M is $0.90M (5.9%) behind budget."),
    ("Margin","EBITDA of $0.36M gave a 14.4% margin against a 17.4% budget. YTD EBITDA of $2.16M is $0.44M behind plan; the gap is driven by revenue, not cost overrun."),
    ("Costs","Cost control is holding — total costs ran at 97.7% of budget for the month and 96.4% YTD. The underspend has absorbed a meaningful share of the revenue shortfall."),
    ("Collections","The June collection rate of 96.0% is the strongest month of the year so far. AR stands at $6.70M, of which $1.90M (28.4%) is now 60+ days overdue."),
    ("Working capital","Net working capital has absorbed $1.40M of cash since 1 January, almost entirely through the AR build. Payables have been drawn down $0.20M over the same period."),
    ("Cash","Cash closed at $4.08M, down $0.72M since 1 January. At the trailing three-month burn of $0.14M per month that is 29.1 months of runway, before the $1.50M current portion of Facility A."),
], 1):
    put(ws, r, 1, i, kind="plain", halign="center")
    put(ws, r, 2, topic, kind="input")
    put(ws, r, 3, text,  kind="input", span=11, wrap=True)
    ws.row_dimensions[r].height = 32
    r += 1

ws.sheet_view.showGridLines = False
wb.save(args.out)
print(f"✓ Created {args.out}")
print(f"  Tabs: {', '.join(wb.sheetnames)}  ·  {r} rows  ·  active month = {ACTIVE} ({MONTHS[ACTIVE-1]})")
